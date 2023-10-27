import time
import logging
import openpyxl

from aiogram import types, F, Router, Bot, Dispatcher
from aiogram.handlers import message
from aiogram.types import Message, Update
from aiogram.filters import Command, CommandStart
from aiogram.utils.keyboard import InlineKeyboardBuilder, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardBuilder
import aiogram.filters.callback_data as filters
# from aiogram.contrib.fsm_storage.memory import MemoryStorage # Для проф бота
from aiogram.utils.markdown import hbold
from aiogram.fsm.context import FSMContext
from aiogram.filters.state import State, StatesGroup
from aiogram.enums import ParseMode
import sqlite3
import os


TOKEN = os.environ['TOKEN']
router = Router()
bot = Bot(token=TOKEN,parse_mode=ParseMode.HTML)


# Создание объекта бота Aiogram
bot = Bot(token="YOUR_TELEGRAM_BOT_TOKEN")
dp = Dispatcher(bot)


# Обработчик команды /import
@dp.message(commands=['import'])
async def import_excel(message: Message):
    file_id = message.reply_to_message.document.file_id
    file = await bot.get_file(file_id)
    file_path = "C:/Users/Sokol/Documents/GitHub/Hackaton-Bot/data_read/Schedule.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook.active
    sheet = workbook["Sheet1"]

    cell_value = sheet["A1"].value
    cell_value = sheet.cell(row=1, column=1).value

    for row in sheet.iter_rows(min_row=1, max_row=10,values_only=True)
    await message.anwer(f'{row}')


